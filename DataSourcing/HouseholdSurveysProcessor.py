import os
import glob
import re
from openpyxl import load_workbook
import pandas as pd
import pprint
import numpy

class HouseholdSurveysProcessor:

    def __init__(self, file_storage):
        self._survey_data_folder = 'survey_data'
        self._input_base_path = f'{file_storage.get_raw_base_path()}/{self._survey_data_folder}/'
        self._food_sufficiency_filter = 'food sufficiency'
        self._children_information_filter = 'with children'
        self._characteristics = [{'topic': 'total', 'characteristic': 'total', 'exact_match': True},
                                 {'topic': 'age', 'characteristic': '18 - 24', 'exact_match': True},
                                 {'topic': 'age', 'characteristic': '25 - 39', 'exact_match': True},
                                 {'topic': 'age', 'characteristic': '40 - 54', 'exact_match': True},
                                 {'topic': 'age', 'characteristic': '55 - 64', 'exact_match': True},
                                 {'topic': 'age', 'characteristic': '65 and above', 'exact_match': True},
                                 {'topic': 'gender', 'characteristic': 'male', 'exact_match': True},
                                 {'topic': 'gender', 'characteristic': 'female', 'exact_match': True},
                                 {'topic': 'race', 'characteristic': 'hispanic or latino', 'exact_match': False},
                                 {'topic': 'race', 'characteristic': 'white alone', 'exact_match': False},
                                 {'topic': 'race', 'characteristic': 'black alone', 'exact_match': False},
                                 {'topic': 'race', 'characteristic': 'asian alone', 'exact_match': False},
                                 {'topic': 'race', 'characteristic': 'two or more races', 'exact_match': False}]
        self._columns = ['State',
                         'Topic',
                         'Characteristic',
                         'Total',
                         'Enough of the kinds of food wanted',
                         'Enough food but not always the kinds wanted',
                         'Sometimes not enough to eat',
                         'Often not enough to eat',
                         'Did not report']
        self._processed_survey_data = [
            {'input_folder': '/survey_data/standard/all/', 'output_folder': '/consolidated_survey_data/standard/all/', 'normalized_data': True},
            {'input_folder': '/survey_data/standard/children/', 'output_folder': '/consolidated_survey_data/standard/children/', 'normalized_data': True},
            {'input_folder': '/survey_data/errors/all/', 'output_folder': '/consolidated_survey_data/errors/all/', 'normalized_data': False},
            {'input_folder': '/survey_data/errors/children/', 'output_folder': '/consolidated_survey_data/errors/children/', 'normalized_data': False}
        ]
        # self._characteristics = ['hispanic or latino']
        self._file_storage = file_storage

    def process_survey_data(self):
        survey_files = self.retrieve_survey_files()
        print('Survey Files', survey_files)
        for survey_file in survey_files:
            # survey_file = 'raw_data/survey_data/2020/wk1/food3b_week1.xlsx'
            wb = load_workbook(survey_file)
            print('Reading the workbook and processing into a csv', survey_file)
            if self.__workbook_contains_food_sufficiency(wb) and not self.__workbook_contains_prior_to_covid_19(wb):
                print('Contains food sufficiency information', survey_file)

                contains_errors = self.__file_contains_standard_errors(survey_file)
                print('Does it contain standard errors?', contains_errors)
                contains_child_info = self.workbook_includes_information_on_children(wb)
                print('Does it include information on children?', contains_child_info)

                self.__process_workbook(wb, contains_errors=contains_errors, contains_child_info=contains_child_info)

        wb.close()

    def consolidate_survey_data(self):
        for survey_data in self._processed_survey_data:
            processed_data_folder = self._file_storage.get_processed_base_path() + survey_data['input_folder']
            normalized_data = survey_data['normalized_data']
            processed_files = list(glob.iglob(processed_data_folder + '*.csv'))
            consolidated_data = {}
            consolidated_normalized_data = {}
            print('Processing files in folder', processed_data_folder)
            print('Processing files', processed_files)
            for file in processed_files:
                print('Processing', file)
                extracted_week = os.path.basename(file).replace('Week', '').replace('.csv', '').strip()
                weekly_data = pd.read_csv(file, index_col=False)
                for group in weekly_data.groupby(['State']):
                    state = group[0]
                    frame = group[1]
                    frame['Week'] = extracted_week
                    for column in self._columns[3:]:
                        frame[column] = frame[column].apply(lambda x: 0 if str(x).strip() == '-' else numpy.float32(x))

                    if normalized_data:
                        normalized_frame = frame.copy(deep=True)
                        total_row = frame[frame['Characteristic'] == 'total']
                        total = numpy.float32(total_row['Total'].values[0])
                        for column in self._columns[3:]:
                            normalized_frame[column] = normalized_frame[column].apply(lambda x: x / total)

                    if state in consolidated_data:
                        consolidated_data[state] = consolidated_data[state].append(frame)
                    else:
                        consolidated_data[state] = frame

                    if normalized_data and state in consolidated_normalized_data:
                        consolidated_normalized_data[state] = consolidated_normalized_data[state].append(normalized_frame)
                    elif normalized_data:
                        consolidated_normalized_data[state] = normalized_frame

            print('Formatting consolidated data into json')
            consolidated_output_path = self._file_storage.get_processed_base_path() + survey_data['output_folder']
            self._file_storage.create_directory_if_not_exists(consolidated_output_path)
            for key in consolidated_data.keys():
                consolidated_data[key].reset_index(drop=True, inplace=True)
                output_file = f'{consolidated_output_path}{key}.csv'
                print('Formatting and saving data to files for state', key)
                print('Output file', output_file)
                consolidated_data[key].to_csv(output_file, index=False)
                if normalized_data:
                    consolidated_output_file = f'{consolidated_output_path}{key}-normalized.csv'
                    consolidated_normalized_data[key].reset_index(drop=True, inplace=True)
                    consolidated_normalized_data[key].to_csv(consolidated_output_file, index=False)


    def retrieve_survey_files(self):
        return list(glob.iglob(self._input_base_path + '**/*.xlsx', recursive=True))

    def __workbook_contains_food_sufficiency(self, wb):
        first_cell = wb['US']['A1'].value
        return self._food_sufficiency_filter in str(first_cell).lower()

    def __workbook_contains_prior_to_covid_19(self, wb):
        first_cell = wb['US']['A1'].value
        return 'prior to covid-19' in str(first_cell).lower()

    def __file_contains_standard_errors(self, filepath):
        return '_se_' in filepath

    def workbook_includes_information_on_children(self, wb):
        first_cell = wb['US']['A1'].value
        return self._children_information_filter in str(first_cell).lower()

    def __process_workbook(self, wb, contains_errors, contains_child_info):
        sheet_names = wb.sheetnames
        r = re.compile(r'[A-Z]{2}')
        filtered_sheet_names = list(filter(r.match, sheet_names))
        week_information = wb['US']['A2'].value
        week_regex = re.compile(r'Week [0-9]{1,2}')
        week = week_regex.search(week_information).group(0)

        output = ','.join(self._columns) + '\n'
        for sheet_name in filtered_sheet_names:
            if sheet_name != 'US':
                for c in self._characteristics:
                    sheet = wb[sheet_name]
                    characteristic = c['characteristic']
                    topic = c['topic']
                    exact_match = c['exact_match']
                    output += f'{sheet_name},{topic},{characteristic},'
                    for row in sheet.iter_rows():
                        if row[0].value is not None:
                            row_title = row[0].value.lower().strip()
                            if row_title == characteristic or (not exact_match and row_title.startswith(characteristic)):
                                output += ','.join([str(statistic.value) for statistic in row[1:]]) + '\n'
                                break



        output_path = '/'.join([
            self._survey_data_folder,
            'errors' if contains_errors else 'standard',
            'children' if contains_child_info else 'all',
            f'{week}.csv'])

        print(output_path)
        self._file_storage.store_as_processed_file(output_path, output)


if __name__ == '__main__':
    from FileStorage import FileStorage
    HouseholdSurveysProcessor(FileStorage()).consolidate_survey_data()
