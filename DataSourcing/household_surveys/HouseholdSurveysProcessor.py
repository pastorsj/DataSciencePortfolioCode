import os
import glob
import re
from openpyxl import load_workbook
import pandas as pd
import numpy
import json
import S3Api

# Constants
STORE_DATA = True

class HouseholdSurveysProcessor:
    """Processes household survey data in csv files"""

    def __init__(self, file_storage, s3_api):
        """ Create a new instance of the HouseholdSurveysProcessor class

        Parameters
        ----------
        :param file_storage: FileStorage, Required
            The file storage class used to store raw/processed data
        :param s3_api: S3_API, Required
            The S3 api wrapper class used to store data in AWS S3

        ----------
        """
        self._survey_data_folder = 'survey_data'
        self._input_base_path = f'{file_storage.get_raw_base_path()}/{self._survey_data_folder}/'
        # Some filters for the raw xlsx files
        self._food_sufficiency_filter = 'food sufficiency'
        self._children_information_filter = 'with children'
        # The desired breakdown characteristics from each survey
        self._characteristics = [{'topic': 'total', 'characteristic': 'total', 'exact_match': True},
                                 {'topic': 'age', 'characteristic': '18 - 24', 'exact_match': True},
                                 {'topic': 'age', 'characteristic': '25 - 39', 'exact_match': True},
                                 {'topic': 'age', 'characteristic': '40 - 54', 'exact_match': True},
                                 {'topic': 'age', 'characteristic': '55 - 64', 'exact_match': True},
                                 {'topic': 'age', 'characteristic': '65 and above', 'exact_match': True},
                                 {'topic': 'gender', 'characteristic': 'male', 'exact_match': True},
                                 {'topic': 'gender', 'characteristic': 'female', 'exact_match': True},
                                 {'topic': 'race', 'characteristic': 'hispanic or latino', 'exact_match': False},
                                 {'topic': 'race', 'characteristic': 'white alone', 'exact_match': False},
                                 {'topic': 'race', 'characteristic': 'black alone', 'exact_match': False},
                                 {'topic': 'race', 'characteristic': 'asian alone', 'exact_match': False},
                                 {'topic': 'race', 'characteristic': 'two or more races', 'exact_match': False}]
        # The list of columns for each csv
        self._columns = ['State',
                         'Topic',
                         'Characteristic',
                         'Total',
                         'Enough of the kinds of food wanted',
                         'Enough food but not always the kinds wanted',
                         'Sometimes not enough to eat',
                         'Often not enough to eat',
                         'Did not report']
        # The map of processed data folders
        self._processed_survey_data = [
            {'input_folder': '/survey_data/standard/all/', 'output_folder': '/consolidated_survey_data/standard/all/',
             'normalized_data': True},
            {'input_folder': '/survey_data/standard/children/',
             'output_folder': '/consolidated_survey_data/standard/children/', 'normalized_data': True},
            {'input_folder': '/survey_data/errors/all/', 'output_folder': '/consolidated_survey_data/errors/all/',
             'normalized_data': False},
            {'input_folder': '/survey_data/errors/children/',
             'output_folder': '/consolidated_survey_data/errors/children/', 'normalized_data': False}
        ]
        self._file_storage = file_storage
        self._s3_api = s3_api

    def process_survey_data(self):
        """Processes each survey file"""
        survey_files = self.retrieve_survey_files()
        print('Survey Files', survey_files)
        for survey_file in survey_files:
            wb = load_workbook(survey_file)
            print('Reading the workbook and processing into a csv', survey_file)
            # If the file contains information on food sufficiency and does not contain data prior to Covid19,
            # continue processing the file
            if self.__workbook_contains_food_sufficiency(wb) and not self.__workbook_contains_prior_to_covid_19(wb):
                print('Contains food sufficiency information', survey_file)

                # Survey files sometimes contain information on standard errors
                contains_errors = self.__file_contains_standard_errors(survey_file)
                print('Does it contain standard errors?', contains_errors)
                # Survey files sometimes contain information on families with children versus without
                contains_child_info = self.workbook_includes_information_on_children(wb)
                print('Does it include information on children?', contains_child_info)

                self.__process_workbook(wb, contains_errors=contains_errors, contains_child_info=contains_child_info)

        wb.close()

    def consolidate_survey_data(self):
        """
        Consolidate the survey data by grouped the initially processed survey files by state and
        normalizing the population answering the survey questions
        """
        survey_metadata = json.load(open('raw_data/survey_data/survey_metadata.json', 'r'))
        for survey_data in self._processed_survey_data:
            processed_data_folder = self._file_storage.get_processed_base_path() + survey_data['input_folder']
            # In some cases, we are not normalizing the data (for standard errors)
            normalized_data = survey_data['normalized_data']
            processed_files = list(glob.iglob(processed_data_folder + '*.csv'))
            consolidated_data = {}
            consolidated_normalized_data = {}
            print('Processing files in folder', processed_data_folder)
            print('Processing files', processed_files)
            for file in processed_files:
                print('Processing', file)
                extracted_week = os.path.basename(file).replace('Week', '').replace('.csv', '').strip()
                weekly_data = pd.read_csv(file, index_col=False)
                for group in weekly_data.groupby(['State']):
                    state = group[0]
                    frame = group[1]
                    frame['Week'] = extracted_week
                    frame['Date'] = survey_metadata[extracted_week]
                    for column in self._columns[3:]:
                        frame[column] = frame[column].apply(lambda x: 0 if str(x).strip() == '-' else numpy.float32(x))

                    if normalized_data:
                        normalized_frame = frame.copy(deep=True)
                        total_row = frame[frame['Characteristic'] == 'total']
                        total = numpy.float32(total_row['Total'].values[0])
                        # Normalize the totals
                        for column in self._columns[3:]:
                            normalized_frame[column] = normalized_frame[column].apply(lambda x: x / total)

                    if state in consolidated_data:
                        consolidated_data[state] = consolidated_data[state].append(frame)
                    else:
                        consolidated_data[state] = frame

                    if normalized_data and state in consolidated_normalized_data:
                        consolidated_normalized_data[state] = consolidated_normalized_data[state].append(
                            normalized_frame)
                    elif normalized_data:
                        consolidated_normalized_data[state] = normalized_frame

            print('Formatting consolidated data into json')
            consolidated_output_path = self._file_storage.get_processed_base_path() + survey_data['output_folder']
            self._file_storage.create_directory_if_not_exists(consolidated_output_path)
            for key in consolidated_data.keys():
                consolidated_data[key].reset_index(drop=True, inplace=True)
                output_file = f'{consolidated_output_path}{key}.csv'
                print('Formatting and saving data to files for state', key)
                print('Output file', output_file)
                consolidated_data[key].to_csv(output_file, index=False)
                if normalized_data:
                    consolidated_output_file = f'{consolidated_output_path}{key}-normalized.csv'
                    consolidated_normalized_data[key].reset_index(drop=True, inplace=True)
                    consolidated_normalized_data[key].to_csv(consolidated_output_file, index=False)

    def retrieve_survey_files(self):
        """Retrieve all raw survey files"""
        return list(glob.iglob(self._input_base_path + '**/*.xlsx', recursive=True))

    def __workbook_contains_food_sufficiency(self, wb):
        """Determines whether the workbook contains information on food sufficiency

        Parameters
        ----------
        :param wb: openpyxl.Workbook, Required
            The data structure containing the excel workbook

        ----------
        """
        first_cell = wb['US']['A1'].value
        return self._food_sufficiency_filter in str(first_cell).lower()

    def __workbook_contains_prior_to_covid_19(self, wb):
        """Determines whether the workbook contains information prior to Covid19

        Parameters
        ----------
        :param wb: openpyxl.Workbook, Required
            The data structure containing the excel workbook

        ----------
        """
        first_cell = wb['US']['A1'].value
        return 'prior to covid-19' in str(first_cell).lower()

    def __file_contains_standard_errors(self, file_path):
        """Determines whether the file contains standard errors

        Parameters
        ----------
        :param file_path: String, Required
            The path of the file

        ----------
        """
        return '_se_' in file_path

    def workbook_includes_information_on_children(self, wb):
        """Determines whether the file contains information on children

        Parameters
        ----------
        :param wb: openpyxl.Workbook, Required
            The data structure containing the excel workbook

        ----------
        """
        first_cell = wb['US']['A1'].value
        return self._children_information_filter in str(first_cell).lower()

    def __process_workbook(self, wb, contains_errors, contains_child_info):
        """Processes each individual workbook

        Parameters
        ----------
        :param wb: openpyxl.Workbook, Required
            The data structure containing the excel workbook
        :param contains_errors: Boolean, Required
            Whether the workbook contains standard errors rather than normal values
        :param contains_child_info: Boolean, Required
            Whether the workbook contains informatin on families with children

        ----------
        """
        sheet_names = wb.sheetnames
        r = re.compile(r'[A-Z]{2}')
        filtered_sheet_names = list(filter(r.match, sheet_names))
        week_information = wb['US']['A2'].value
        week_regex = re.compile(r'Week [0-9]{1,2}')
        week = week_regex.search(week_information).group(0)

        output = ','.join(self._columns) + '\n'
        for sheet_name in filtered_sheet_names:
            if sheet_name != 'US':
                for c in self._characteristics:
                    sheet = wb[sheet_name]
                    characteristic = c['characteristic']
                    topic = c['topic']
                    exact_match = c['exact_match']
                    output += f'{sheet_name},{topic},{characteristic},'
                    for row in sheet.iter_rows():
                        if row[0].value is not None:
                            row_title = row[0].value.lower().strip()
                            if row_title == characteristic or (
                                    not exact_match and row_title.startswith(characteristic)):
                                output += ','.join([str(statistic.value) for statistic in row[1:]]) + '\n'
                                break

        output_path = '/'.join([
            self._survey_data_folder,
            'errors' if contains_errors else 'standard',
            'children' if contains_child_info else 'all',
            f'{week}.csv'])

        print(output_path)
        self._file_storage.store_as_processed_file(output_path, output)

    def store_survey_data(self):
        print('Store processed survey data in S3')

        processed_files = list(glob.iglob('processed_data/consolidated_survey_data/**/*.csv', recursive=True))
        for file in processed_files:
            print('Opening file', file)
            contents = open(file, 'rb')
            print('Uploading', file, 'to S3')
            self._s3_api.upload_bytes(contents, file.replace('processed_data/', ''), S3Api.S3Location.PROCESSED_DATA)
            contents.close()

        print('Uploaded all files')


if __name__ == '__main__':
    from dotenv import load_dotenv
    from FileStorage import FileStorage
    load_dotenv()

    household_surveys_processor_instance = HouseholdSurveysProcessor(FileStorage(), S3Api.S3Api())

    # print('Processing survey data')
    # household_surveys_processor_instance.process_survey_data()
    # print('Consolidating survey data')
    # household_surveys_processor_instance.consolidate_survey_data()

    if STORE_DATA:
        print('Storing processed survey data in S3')
        household_surveys_processor_instance.store_survey_data()
